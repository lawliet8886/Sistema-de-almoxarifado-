import sys
import sqlite3
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QMessageBox, QTableWidget, QTableWidgetItem, QComboBox, QDialog, QFormLayout, QAbstractItemView, QDateEdit, QFileDialog
from PyQt5.QtGui import QColor, QBrush, QIntValidator
from PyQt5.QtCore import Qt, QDate, QTimer
from PyQt5 import QtGui
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import unidecode
import datetime
import re
from openpyxl.drawing.image import Image

# Conexão com o banco de dados
conn = sqlite3.connect('almoxarifado.db')
c = conn.cursor()

# Criação da tabela de produtos no banco de dados
c.execute('''CREATE TABLE IF NOT EXISTS produtos (
                nome TEXT NOT NULL,
                categoria TEXT NOT NULL,
                quantidade INTEGER NOT NULL,
                quantidade_minima INTEGER NOT NULL,
                vencimento TEXT NOT NULL,
                data_registro TEXT NOT NULL,
                retirada TEXT,
                valor_unitario REAL,
                valor_total REAL
            )''')

# Criação da tabela de retiradas no banco de dados
c.execute('''CREATE TABLE IF NOT EXISTS retiradas (
                produto TEXT NOT NULL,
                quantidade INTEGER NOT NULL,
                valor_unitario REAL,
                valor_total REAL,
                data TEXT NOT NULL,
                responsavel TEXT NOT NULL
            )''')
            
# Criação da tabela de entradas no banco de dados
c.execute('''CREATE TABLE IF NOT EXISTS entradas (
                produto TEXT NOT NULL,
                quantidade INTEGER NOT NULL,
                valor_unitario REAL,
                valor_total REAL,
                data TEXT NOT NULL,
                responsavel TEXT NOT NULL
            )''')
                                   
class EditDialog(QDialog):
    def __init__(self, produto, main_window, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Editar Item")
        self.layout = QVBoxLayout(self)
        self.produto = produto
        self.main_window = main_window

        self.form_layout = QFormLayout()
        self.layout.addLayout(self.form_layout)

        self.name_input = QLineEdit()
        self.name_input.setText(produto[0])
        self.form_layout.addRow("Nome:", self.name_input)

        self.category_combo = QComboBox()
        self.category_combo.addItem("Itens de Escritório")
        self.category_combo.addItem("Itens Médicos")
        self.category_combo.addItem("Itens de Limpeza")
        self.category_combo.addItem("medicações")
        
        self.form_layout.addRow("Categoria:", self.category_combo)

        self.quantity_input = QLineEdit()
        self.quantity_input.setText(str(produto[2]))
        self.form_layout.addRow("Quantidade:", self.quantity_input)

        self.min_quantity_input = QLineEdit()
        self.min_quantity_input.setText(str(produto[3]))
        self.form_layout.addRow("Quantidade Mínima:", self.min_quantity_input)

        self.expiry_date_layout = QHBoxLayout()
        self.form_layout.addRow("Data de Vencimento:", self.expiry_date_layout)

        self.day_label = QLabel("Dia (opcional):")
        self.day_input = QLineEdit()
        self.expiry_date_layout.addWidget(self.day_label)
        self.expiry_date_layout.addWidget(self.day_input)

        self.month_label = QLabel("Mês:")
        self.month_input = QLineEdit()
        self.expiry_date_layout.addWidget(self.month_label)
        self.expiry_date_layout.addWidget(self.month_input)

        self.year_label = QLabel("Ano:")
        self.year_input = QLineEdit()
        self.expiry_date_layout.addWidget(self.year_label)
        self.expiry_date_layout.addWidget(self.year_input)

        self.register_date_layout = QHBoxLayout()
        self.form_layout.addRow("Data de Registro:", self.register_date_layout)

        self.unit_value_input = QLineEdit()
        self.unit_value_input.setText(str(produto[7]))
        self.form_layout.addRow("Valor Unitário:",  self.unit_value_input)

        self.register_day_label = QLabel("Dia:")
        self.register_day_input = QLineEdit()
        self.register_date_layout.addWidget(self.register_day_label)
        self.register_date_layout.addWidget(self.register_day_input)

        self.register_month_label = QLabel("Mês:")
        self.register_month_input = QLineEdit()
        self.register_date_layout.addWidget(self.register_month_label)
        self.register_date_layout.addWidget(self.register_month_input)

        self.register_year_label = QLabel("Ano:")
        self.register_year_input = QLineEdit()
        self.register_date_layout.addWidget(self.register_year_label)
        self.register_date_layout.addWidget(self.register_year_input)

        self.button_layout = QHBoxLayout()
        self.layout.addLayout(self.button_layout)

        self.save_button = QPushButton("Salvar")
        self.save_button.clicked.connect(self.save_changes)
        self.button_layout.addWidget(self.save_button)

        self.delete_button = QPushButton("Excluir")
        self.delete_button.clicked.connect(self.delete_product)
        self.button_layout.addWidget(self.delete_button)
        self.total_value_label = QLabel()

        # Preencher campos de dia, mês e ano corretamente
        vencimento = produto[4].split("/")
        if len(vencimento) > 0:
            self.day_input.setText(vencimento[0])
        if len(vencimento) > 1:
            self.month_input.setText(vencimento[1])
        if len(vencimento) > 2:
            self.year_input.setText(vencimento[2])

        data_registro = produto[6].split("/")
        if len(data_registro) > 0:
            self.register_day_input.setText(data_registro[0])
        if len(data_registro) > 1:
            self.register_month_input.setText(data_registro[1])
        if len(data_registro) > 2:
            self.register_year_input.setText(data_registro[2])

    def save_changes(self):
        nome = self.name_input.text()
        categoria = self.category_combo.currentText()
        quantidade = self.quantity_input.text()
        min_quantidade = self.min_quantity_input.text()
        dia_vencimento = self.day_input.text()
        mes_vencimento = self.month_input.text()
        ano_vencimento = self.year_input.text()
        dia_registro = self.register_day_input.text()
        mes_registro = self.register_month_input.text()
        ano_registro = self.register_year_input.text()
        valor_unitario = self.unit_value_input.text()
        

        if nome and categoria and quantidade and min_quantidade and  mes_vencimento and ano_vencimento and dia_registro and mes_registro and ano_registro and valor_unitario:
            try:
                quantidade = int(quantidade)
                min_quantidade = int(min_quantidade)
                valor_unitario = float(valor_unitario)
                valor_total = quantidade * valor_unitario 
                vencimento = ""
                if not dia_vencimento:
                    dia_vencimento = "01"
                vencimento += f"{dia_vencimento}/"
                vencimento += f"{mes_vencimento}/{ano_vencimento}"
                data_registro = ""
                if not dia_registro:
                    dia_registro = "01"
                data_registro += f"{dia_registro}/"
                data_registro += f"{mes_registro}/{ano_registro}"
                c.execute("UPDATE produtos SET nome=?, categoria=?, quantidade=?, quantidade_minima=?, vencimento=?, data_registro=?, valor_unitario=?, valor_total=? WHERE nome=?",
                          (nome, categoria, quantidade, min_quantidade, vencimento, data_registro, valor_unitario, valor_total, self.produto[0]))
                conn.commit()
                QMessageBox.information(self, "Sucesso", "Alterações salvas com sucesso!")
                self.main_window.load_products()
                self.accept()

            except ValueError:
                QMessageBox.warning(self, "Erro", "Quantidade inválida. Insira um valor numérico.")
        else:
            QMessageBox.warning(self, "Erro", "Preencha todos os campos obrigatórios.")

    def delete_product(self):
        reply = QMessageBox.question(self, "Excluir Item", "Tem certeza que deseja excluir este item?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            c.execute("DELETE FROM produtos WHERE nome=?", (self.produto[0],))
            c.execute("DELETE FROM retiradas WHERE produto=?", (self.produto[0],))
            conn.commit()
            QMessageBox.information(self, "Sucesso", "Item excluído com sucesso!")
            self.accept()
            self.main_window.update_total_value()
            
                                    
class WithdrawalDialog(QDialog):
    def __init__(self, quantidade_atual, valor_unitario, produto, parent=None):
        super(WithdrawalDialog, self).__init__(parent)
        self.setWindowTitle("Registrar Retirada")
        self.setFixedSize(655, 500)  # Define o tamanho fixo da janela
        self.layout = QVBoxLayout(self)
        self.produto = produto
        self.valor_unitario = valor_unitario
        self.quantity_label = QLabel("Quantidade:")
        self.layout.addWidget(self.quantity_label)

        self.quantity_input = QLineEdit()
        self.layout.addWidget(self.quantity_input)

        self.date_label = QLabel("Data:")
        self.layout.addWidget(self.date_label)

        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        self.layout.addWidget(self.date_edit)

        self.responsavel_label = QLabel("Responsável:")
        self.layout.addWidget(self.responsavel_label)

        self.responsavel_input = QLineEdit()
        self.layout.addWidget(self.responsavel_input)

        self.button_layout = QHBoxLayout()
        self.layout.addLayout(self.button_layout)

        self.save_button = QPushButton("Salvar")
        self.save_button.clicked.connect(self.accept)
        self.button_layout.addWidget(self.save_button)

        self.quantity_maxima = quantidade_atual
        self.quantity_input.setValidator(QtGui.QIntValidator(1, quantidade_atual, self))

    def accept(self):
        quantidade_retirada = self.quantity_input.text()
        if quantidade_retirada.isdigit() and int(quantidade_retirada) > 0:
            super().accept()
        else:
            QMessageBox.warning(self, "Erro", "A quantidade de retirada deve ser um valor inteiro maior que zero.")

            
class EntryDialog(QDialog):
    def __init__(self, quantidade_maxima, valor_unitario, produto, parent=None):
        super(EntryDialog, self).__init__(parent)
        self.setWindowTitle("Registrar Entrada")
        self.setFixedSize(655, 500)  # Define o tamanho fixo da janela
        self.layout = QVBoxLayout(self)
        self.produto = produto
        self.valor_unitario = valor_unitario
        self.quantity_label = QLabel("Quantidade:")
        self.layout.addWidget(self.quantity_label)

        self.quantity_input = QLineEdit()
        self.layout.addWidget(self.quantity_input)

        self.date_label = QLabel("Data:")
        self.layout.addWidget(self.date_label)

        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        self.layout.addWidget(self.date_edit)

        self.responsavel_label = QLabel("Responsável:")
        self.layout.addWidget(self.responsavel_label)

        self.responsavel_input = QLineEdit()
        self.layout.addWidget(self.responsavel_input)

        self.button_layout = QHBoxLayout()
        self.layout.addLayout(self.button_layout)

        self.save_button = QPushButton("Salvar")
        self.save_button.clicked.connect(self.accept)
        self.button_layout.addWidget(self.save_button)

        self.quantity_input.setValidator(QtGui.QIntValidator(1, 99999, self))

    def accept(self):
        quantidade_entrada = self.quantity_input.text()
        if quantidade_entrada.isdigit() and int(quantidade_entrada) > 0:
            super().accept()
        else:
            QMessageBox.warning(self, "Erro", "A quantidade de entrada deve ser um valor inteiro maior que zero.")


         
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Sistema de Gerenciamento de Almoxarifado")
        self.setGeometry(100, 100, 600, 400)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        self.layout = QVBoxLayout()
        self.central_widget.setLayout(self.layout)

        self.label = QLabel("Nome do item para adicionar:")
        self.layout.addWidget(self.label)

        self.name_input = QLineEdit()
        self.layout.addWidget(self.name_input)

        self.category_label = QLabel("Categoria:")
        self.layout.addWidget(self.category_label)

        self.category_combo = QComboBox()
        self.category_combo.addItem("Itens de Escritório")
        self.category_combo.addItem("Itens Médicos")
        self.category_combo.addItem("Itens de Limpeza")  
        self.category_combo.addItem("Medicações")
        self.category_combo.addItem("Itens Importados")
                      
        self.layout.addWidget(self.category_combo)

        self.quantity_label = QLabel("Quantidade:")
        self.layout.addWidget(self.quantity_label)

        self.quantity_input = QLineEdit()
        self.layout.addWidget(self.quantity_input)

        self.min_quantity_label = QLabel("Quantidade Mínima:")
        self.layout.addWidget(self.min_quantity_label)

        self.min_quantity_input = QLineEdit()
        self.layout.addWidget(self.min_quantity_input)
        
        self.expiry_date_layout = QHBoxLayout()
        self.layout.addLayout(self.expiry_date_layout)

        self.expiry_date_label = QLabel("Data de Vencimento:")
        self.expiry_date_layout.addWidget(self.expiry_date_label)

        self.day_label = QLabel("Dia (opcional):")
        self.day_input = QLineEdit()
        
        self.expiry_date_layout.addWidget(self.day_label)
        self.expiry_date_layout.addWidget(self.day_input)
        self.month_label = QLabel("Mês:")
        self.month_input = QLineEdit()
        self.month_input.textChanged.connect(self.check_month_input)
        self.expiry_date_layout.addWidget(self.month_label)
        self.expiry_date_layout.addWidget(self.month_input)        
        # Conectar o sinal `returnPressed` com o slot `focusNextChild`.
        self.name_input.returnPressed.connect(self.focusNextChild)
        self.quantity_input.returnPressed.connect(self.focusNextChild)
        self.min_quantity_input.returnPressed.connect(self.focusNextChild)
        self.day_input.returnPressed.connect(self.focusNextChild)
        self.month_input.returnPressed.connect(self.focusNextChild)        
        
        # Configurar a ordem de tabulação.
        self.setTabOrder(self.name_input, self.quantity_input)
        self.setTabOrder(self.quantity_input, self.min_quantity_input)
        self.setTabOrder(self.min_quantity_input, self.day_input)   
        self.setTabOrder(self.day_input, self.month_input)
                              
        self.year_label = QLabel("Ano:")
        self.year_input = QLineEdit()
        self.year_input.textChanged.connect(self.check_year_input)
        self.expiry_date_layout.addWidget(self.year_label)
        self.expiry_date_layout.addWidget(self.year_input)

        self.register_date_layout = QHBoxLayout()
        self.layout.addLayout(self.register_date_layout)

        self.register_date_label = QLabel("Data de Registro:")
        self.register_date_layout.addWidget(self.register_date_label)

        self.unit_value_input = QLineEdit(self)
        self.unit_value_input.   setPlaceholderText("Valor Unitário")
        self.unit_value_input.   setValidator(QIntValidator(0, 999999999, self))
        self.layout.addWidget(self.unit_value_input)

        self.register_day_label = QLabel("Dia:")
        self.register_day_input = QLineEdit()
        self.register_day_input.textChanged.connect(self.check_register_day_input)
        self.register_date_layout.addWidget(self.register_day_label)
        self.register_date_layout.addWidget(self.register_day_input)

        self.register_month_label = QLabel("Mês:")
        self.register_month_input = QLineEdit()
        self.register_month_input.textChanged.connect(self.check_register_month_input)
        self.register_date_layout.addWidget(self.register_month_label)
        self.register_date_layout.addWidget(self.register_month_input)

        self.register_year_label = QLabel("Ano:")
        self.register_year_input = QLineEdit()
        self.register_year_input.textChanged.connect(self.check_register_year_input)
        self.register_date_layout.addWidget(self.register_year_label)
        self.register_date_layout.addWidget(self.register_year_input)
        
        self.add_button = QPushButton("Adicionar Item")
        self.add_button.clicked.connect(self.add_product)
        self.layout.addWidget(self.add_button)

        self.edit_button = QPushButton("Editar Item")
        self.edit_button.clicked.connect(self.edit_product)
        self.layout.addWidget(self.edit_button)

        self.delete_button = QPushButton("Excluir Item")
        self.delete_button.clicked.connect(self.delete_product)
        self.layout.addWidget(self.delete_button)

        self.withdrawal_button = QPushButton("Registrar Retirada")
        self.withdrawal_button.clicked.connect(self.open_withdrawal_dialog)
        self.layout.addWidget(self.withdrawal_button)

        self.entry_button = QPushButton("Registrar Entrada")
        self.entry_button.clicked.connect(self.open_entry_dialog)
        self.layout.addWidget(self.entry_button)
        
        self.report_button = QPushButton("Gerar Relatório")
        self.report_button.clicked.connect(self.generate_report)
        self.layout.addWidget(self.report_button)
        
        self.import_button = QPushButton("Importar")
        self.import_button.clicked.connect(self.import_from_excel)
        self.layout.addWidget(self.import_button)
        self.import_input = QLineEdit()
        
        self.search_button =    QPushButton("Buscar")
        self.search_button.clicked.connect(self.search_products)
        self.layout.addWidget(self.search_button)
        self.search_input = QLineEdit()
        self.layout.addWidget(self.search_input)  
        
        self.total_value_label = QLabel()
        self.layout.addWidget(self.total_value_label)
        
          
        self.category_filter_combo = QComboBox()
        self.category_filter_combo.addItem("Todas as Categorias")
        self.category_filter_combo.addItem("Itens de Escritório")
        self.category_filter_combo.addItem("Itens Médicos")
        self.category_filter_combo.addItem("Itens de Limpeza")
        self.category_filter_combo.addItem("Medicações")
        self.category_filter_combo.addItem("Itens Importados")
              
        self.category_filter_combo.currentIndexChanged.connect(self.load_products)
        self.layout.addWidget(self.category_filter_combo)
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(
            ["Nome", "Categoria", "Quantidade", "Quantidade Mínima", "Data de Vencimento", "Entrada", "Retirada", "Valor Unitário", "Valor Total"])
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)  # Tabela somente leitura
        self.layout.addWidget(self.table)

        self.legend_label = QLabel("Legenda:")
        self.layout.addWidget(self.legend_label)

        self.legend_layout = QHBoxLayout()
        self.layout.addLayout(self.legend_layout)

        self.legend_expired_label = QLabel(" Vencidos")
        self.legend_expired_label.setStyleSheet("background-color: red; color: white; padding: 5px;")
        self.legend_layout.addWidget(self.legend_expired_label)

        self.legend_near_expiry_label = QLabel("Próx. do Venc.")
        self.legend_near_expiry_label.setStyleSheet("background-color: yellow; padding: 5px;")
        self.legend_layout.addWidget(self.legend_near_expiry_label)

        self.legend_valid_label = QLabel(" Válidos")
        self.legend_valid_label.setStyleSheet("background-color: green; color: white; padding: 5px;")
        self.legend_layout.addWidget(self.legend_valid_label)

        self.legend_low_quantity_label = QLabel("Item Acabando")
        self.legend_low_quantity_label.setStyleSheet("background-color: blue; color: white; padding: 5px;")
        self.legend_layout.addWidget(self.legend_low_quantity_label)

        self.legend_zero_quantity_label = QLabel(" Item zerado")
        self.legend_zero_quantity_label.setStyleSheet("background-color: purple; color: white; padding: 5px;")
        self.legend_layout.addWidget(self.legend_zero_quantity_label)

        self.timer = QTimer()
        self.timer.timeout.connect(self.toggle_blink)
        self.timer.start(500)  # Intervalo de 500 ms para alternar o estilo

        self.load_products()
        self.timer_row = -1  # Índice da linha que está piscando
        self.timer_color = QColor("purple")  # Cor roxa para piscar
    def check_month_input(self, text):
        if len(text) == 2:
            self.year_input.setFocus()

    def check_year_input(self, text):
        if len(text) == 4:
            self.register_day_input.setFocus()

    def check_register_day_input(self, text):
        if len(text) == 2:
            self.register_month_input.setFocus()

    def check_register_month_input(self, text):
        if len(text) == 2:
            self.register_year_input.setFocus()

    def check_register_year_input(self, text):
        if len(text) == 4:
            self.unit_value_input.setFocus()                    
                
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            focused_widget = self.focusWidget()
            if isinstance(focused_widget, QPushButton):
                focused_widget.click()
                               
    def import_from_excel(self):
    # Exibir caixa de diálogo com informações sobre o formato do arquivo Excel
        message = ("O arquivo Excel deve ter os seguintes nomes de colunas na terceira linha( não precisam estar nessa ordem; Não precisa estarem todas preenchidas; Não precisa que o arquivo tenha todas essas colunas, pois os valores faltantes serão preenchidos com uma informção padrão do código):\n"
               "Nome, Categoria, Quantidade, Quantidade Mínima, Vencimento, Data do Primeiro Registro, Valor Unitário\n"
               "Deseja continuar com a importação?")
        reply = QMessageBox.question(self, 'Formato do Arquivo Excel', message,  QMessageBox.Yes, QMessageBox.No)
        if reply == QMessageBox.Yes:
        # Continuar com a importação
            options = QFileDialog.Options()
            options |= QFileDialog.ReadOnly
            file_name, _ = QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "",
                                                   "Excel Files (*.xlsx);;All Files (*)", options=options)
            if file_name:
                workbook = openpyxl.load_workbook(file_name)
                sheet = workbook.active
            else:
            	return



    # Obter os nomes das colunas da primeira linha
            column_names = [(cell.value if cell.value is not None else '') for cell in sheet[3]]

            for row in sheet.iter_rows(min_row=4, values_only=True):  # Ignorar a primeira linha (cabeçalho)
        # Criar um dicionário que mapeia os nomes das colunas para os seus respectivos valores
                item_data = {column_name: value for column_name, value in zip(column_names, row)}

        # Obter os valores das colunas, usando valores padrão para as colunas que estão faltando
                item_name = item_data.get("Nome", "Nome Desconhecido")
                categoria = item_data.get("Categoria", "Itens Importados")

                quantidade = item_data. get("Quantidade", 0)
                if quantidade is None or quantidade == '':
                    quantidade = 0
                else:
                	quantidade = float(quantidade)

                quantidade_minima = item_data.get("Quantidade Mínima", 0)
                if quantidade_minima is None or   quantidade_minima == '':
                    quantidade_minima = 0
                else:
                    quantidade_minima = float(quantidade_minima)
 
                vencimento = item_data.get("Vencimento", datetime.datetime(2100, 1, 1))
                if isinstance(vencimento, datetime.datetime):
                    vencimento = vencimento.strftime("%d/%m/%Y")
                if vencimento is None or vencimento == '':
                    vencimento = "01/01/2100"

                entrada = item_data.get("Data do Primeiro Registro", datetime.datetime(2000, 1, 1))
                if isinstance(entrada, datetime.datetime):
                    entrada = entrada.strftime("%d/%m/%Y")
                if entrada is None or entrada == '':
                    entrada = "01/01/2000"

                valor_unitario = item_data.get("Valor Unitário", 0)
                if valor_unitario is None or valor_unitario == '':
                    valor_unitario = 0
                else:
                    valor_unitario = float(valor_unitario)

                
            # Calcular o valor total
                valor_total = quantidade * valor_unitario

                c.execute("INSERT INTO produtos (nome, categoria, quantidade, quantidade_minima, vencimento, data_registro, valor_unitario, valor_total) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                          (item_name, categoria, quantidade, quantidade_minima, vencimento, entrada, valor_unitario, valor_total))
        conn.commit()
        self.load_products()
        self.update_total_value()
 

    def update_total_value(self):
        c.execute("SELECT SUM(valor_total) FROM produtos")
# No método update_total_value
        total_value = c.fetchone()[0]
        if total_value is None:
            total_value = 0
        total_value_str = 'R$ {:,.2f}'.format(total_value).replace('.', '#').replace(',', '.').replace('#', ',')
        self.total_value_label.setText(f"Valor Total do Almoxarifado: {total_value_str}")

                                                                            
    def search_products(self):
        search_text = self.search_input.text()  # Obtém o texto digitado na caixa de busca
        selected_category = self.    category_filter_combo.currentText()

        if selected_category == "Todas as Categorias":
            c.execute("SELECT nome, categoria, quantidade, quantidade_minima, vencimento, retirada, data_registro, valor_unitario, valor_total FROM produtos WHERE nome LIKE ?", (f'%{search_text}%',))
        else:
            c.execute("SELECT nome, categoria, quantidade, quantidade_minima, vencimento, retirada, data_registro, valor_unitario, valor_total FROM produtos WHERE nome LIKE ? AND categoria=?", (f'%{search_text}%', selected_category))

        products = c.fetchall()
        self.update_table(products)
        
    def update_table(self, products):
        self.table.setRowCount(0)

        for row_number, row_data in   enumerate(products):
            self.table.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                item = QTableWidgetItem(str(data))
                self.table.setItem(row_number, column_number, item)

                quantidade = row_data[2]
                valor_unitario = row_data[7]
                valor_total = valor_unitario * quantidade
                item = QTableWidgetItem(str(valor_total))
                self.table.setItem(row_number, 8, item)

                vencimento = row_data[4]
                vencimento_date = QDate.fromString(vencimento, "d/M/yyyy")
                if not vencimento_date.isValid():
                    vencimento_date = QDate.fromString(vencimento, "M/yyyy")
                retirada = row_data[5]
                if retirada:
                    retirada_date = QDate.fromString(retirada, "dd/MM/yyyy")
                today = QDate.currentDate()
                if vencimento_date < today:
                    for column_number in range(self.table.columnCount()):
                        item = self.table.item(row_number, column_number)
                        if item is not None:
                            item.setBackground(QColor("red"))
                            item.setForeground(QColor("white"))
                elif vencimento_date.addMonths(-2) <= today:
                    for column_number in range(self.table.columnCount()):
                        item = self.table.item(row_number, column_number)
                        if item is not None:
                            item.setBackground(QColor("yellow"))
                    else:
                        for column_number in range(self.table.columnCount()):
                            item = self.table.item(row_number, column_number)
                            if item is not None:
                                item.setBackground(QColor("green"))
                                item.setForeground(QColor("white"))

                quantidade = row_data[2]
                quantidade_minima = row_data[3]
                if quantidade == 0:
                    for column_number in range(self.table.columnCount()):
                        item = self.table.item(row_number, column_number)
                        if item is not None:
                            item.setBackground(QColor("purple"))
                            item.setForeground(QColor("white"))
                elif quantidade <= quantidade_minima:
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)

                    if self.timer.isActive():
                        for column_number in range(self.table.columnCount()):
                            item = self.table.item(row_number, column_number)
                            if item is not None:
                                item.setBackground(QColor("blue"))
                                item.setForeground(QColor("white"))
                    else:
                        for column_number in range(self.table.columnCount()):
                            item = self.table.item(row_number, column_number)
                            if item is not None:
                                item.setBackground(QColor("white"))
                                item.setForeground(QColor("blue"))

            button = QPushButton("Detalhes")
            button.clicked.connect(lambda _, p=row_data[0]: self.open_withdrawal_details(p))
            self.table.setCellWidget(row_number, 6, button)

            button = QPushButton("Detalhes")
            button.clicked.connect(lambda _, p=row_data[0]: self.open_entry_details(p))
            self.table.setCellWidget(row_number, 5, button)

    def load_products(self):
        self.table.setRowCount(0)
        selected_category = self.category_filter_combo.currentText()

        if selected_category == "Todas as Categorias":
            c.execute("SELECT nome, categoria, quantidade, quantidade_minima, vencimento, retirada, data_registro, valor_unitario, valor_total FROM produtos")
        else:
            c.execute("SELECT nome, categoria, quantidade, quantidade_minima, vencimento, retirada, data_registro, valor_unitario, valor_total FROM produtos WHERE categoria=?", (selected_category,))

        products = c.fetchall()

        for row_number, row_data in   enumerate(products):
            self.table.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                item = QTableWidgetItem(str(data))
                self.table.setItem(row_number, column_number, item)

                quantidade = row_data[2]
                valor_unitario = row_data[7]
                valor_total = valor_unitario * quantidade
                item = QTableWidgetItem(str(valor_total))
                self.table.setItem(row_number, 8, item)

                vencimento = row_data[4]
                vencimento_date = QDate.fromString(vencimento, "d/M/yyyy")
                if not vencimento_date.isValid():
                    vencimento_date = QDate.fromString(vencimento, "M/yyyy")
                retirada = row_data[5]
                if retirada:
                    retirada_date = QDate.fromString(retirada, "dd/MM/yyyy")
                today = QDate.currentDate()
                if vencimento_date < today:
                    for column_number in range(self.table.columnCount()):
                        item = self.table.item(row_number, column_number)
                        if item is not None:
                            item.setBackground(QColor("red"))
                            item.setForeground(QColor("white"))
                elif vencimento_date.addMonths(-2) <= today:
                    for column_number in range(self.table.columnCount()):
                        item = self.table.item(row_number, column_number)
                        if item is not None:
                            item.setBackground(QColor("yellow"))
                else:
                    for column_number in range(self.table.columnCount()):
                        item = self.table.item(row_number, column_number)
                        if item is not None:
                            item.setBackground(QColor("green"))
                            item.setForeground(QColor("white"))

                quantidade = row_data[2]
                quantidade_minima = row_data[3]
                if quantidade == 0:
                    for column_number in range(self.table.columnCount()):
                        item = self.table.item(row_number, column_number)
                        if item is not None:
                            item.setBackground(QColor("purple"))
                            item.setForeground(QColor("white"))
                elif quantidade <= quantidade_minima:
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)

                    if self.timer.isActive():
                        for column_number in range(self.table.columnCount()):
                            item = self.table.item(row_number, column_number)
                            if item is not None:
                                item.setBackground(QColor("blue"))
                                item.setForeground(QColor("white"))
                    else:
                        for column_number in range(self.table.columnCount()):
                            item = self.table.item(row_number, column_number)
                            if item is not None:
                                item.setBackground(QColor("white"))
                                item.setForeground(QColor("blue"))

            button = QPushButton("Detalhes")
            button.clicked.connect(lambda _, p=row_data[0]: self.open_withdrawal_details(p))
            self.table.setCellWidget(row_number, 6, button)

            button = QPushButton("Detalhes")
            button.clicked.connect(lambda _, p=row_data[0]: self.open_entry_details(p))
            self.table.setCellWidget(row_number, 5, button)

        self.table.resizeColumnsToContents()
        self.update_total_value()

    def toggle_blink(self):
        row = -1
        for row in range(self.table.rowCount()):
            quantidade = int(self.table.item(row, 2).text())
            quantidade_minima = int(self.table.item(row, 3).text())
            if quantidade == 0:
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item.background().color() == QColor("white"):
                        item.setBackground(QColor("purple"))
                        item.setForeground(QColor("white"))
                    else:
                        item.setBackground(QColor("white"))
                        item.setForeground(QColor("purple"))
            elif quantidade <= quantidade_minima:
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item.background().color() == QColor("white"):
                        item.setBackground(QColor("blue"))
                        item.setForeground(QColor("white"))
                    else:
                        item.setBackground(QColor("white"))
                        item.setForeground(QColor("blue"))

        if row == self.timer_row:
            self.timer_row = -1
        else:
            self.timer_row = row

        if self.table.selectedItems():
            self.table.itemSelectionChanged.emit()

    def add_product(self):
        nome = self.name_input.text()
        categoria = self.category_combo.currentText()
        quantidade = self.quantity_input.text()
        min_quantidade = self.min_quantity_input.text()
        dia_vencimento = self.day_input.text()
        mes_vencimento = self.month_input.text()
        ano_vencimento = self.year_input.text()
        dia_registro = self.register_day_input.text()
        mes_registro = self.register_month_input.text()
        ano_registro = self.register_year_input.text()
        valor_unitario = self.unit_value_input.text()
        

        if nome and categoria and quantidade and min_quantidade and  mes_vencimento and ano_vencimento and dia_registro and mes_registro and ano_registro and valor_unitario:
            try:
                quantidade = int(quantidade)
                min_quantidade = int(min_quantidade)
                valor_unitario = int(valor_unitario)
                vencimento = ""
                if not dia_vencimento:
                    dia_vencimento = "01"
                vencimento += f"{dia_vencimento}/"
                vencimento += f"{mes_vencimento}/{ano_vencimento}"
                data_registro = ""
                if not dia_registro:
                    dia_registro = "01"
                data_registro += f"{dia_registro}/"
                data_registro += f"{mes_registro}/{ano_registro}"
                c.execute("INSERT INTO produtos VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                          (nome, categoria, quantidade, min_quantidade, vencimento, data_registro, None, valor_unitario, valor_unitario*quantidade))                                                   
                if quantidade <= 0:
                      self.timer_row = self.table.rowCount() - 1

                conn.commit()
                QMessageBox.information(self, "Sucesso", "Item adicionado com sucesso!")
                self.clear_input_fields()
                self.update_total_value()
                self.load_products()
            except ValueError:
                QMessageBox.warning(self, "Erro", "Quantidade inválida. Insira um valor numérico.")
        else:
            QMessageBox.warning(self, "Erro", "Preencha todos os campos obrigatórios.")
            
            

    def edit_product(self):
        selected_items = self.table.selectedItems()
        if len(selected_items) == 0:
            QMessageBox.warning(self, "Erro", "Selecione um item para editar.")
            return

        selected_row = selected_items[0].row()
        nome = self.table.item(selected_row, 0).text()
        categoria = self.table.item(selected_row, 1).text()
        quantidade = self.table.item(selected_row, 2).text()
        min_quantidade = self.table.item(selected_row, 3).text()
        vencimento = self.table.item(selected_row, 4).text()
        data_registro = self.table.item(selected_row, 6).text()
        valor_unitario = self.table.item(selected_row, 7).text()

        produto = (nome, categoria, quantidade, min_quantidade, vencimento, None, data_registro, valor_unitario)

        dialog = EditDialog(produto, self)
        if dialog.exec_() == QDialog.Accepted:
            self.load_products()

    def delete_product(self):
        selected_items = self.table.selectedItems()
        if len(selected_items) == 0:
            QMessageBox.warning(self, "Erro", "Selecione um item para excluir.")
            return

        selected_row = selected_items[0].row()
        nome = self.table.item(selected_row, 0).text()

        reply = QMessageBox.question(self, "Excluir Item", "Tem certeza que deseja excluir este item?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            c.execute("DELETE FROM produtos WHERE nome=?", (nome,))
            c.execute("DELETE FROM retiradas WHERE produto=?", (nome,))
            conn.commit()
            QMessageBox.information(self, "Sucesso", "Item excluído com sucesso!")
            self.load_products()            
            self.update_total_value()

    def open_withdrawal_dialog(self):
        selected_items = self.table.selectedItems()
        if len(selected_items) == 0:
            QMessageBox.warning(self, "Erro", "Selecione um item para registrar a retirada.")
            return

        selected_row = selected_items[0].row()
        quantidade_atual = int(self.table.item(selected_row, 2).text())
        valor_unitario = float(self.table.item(selected_row, 7).text())
        produto = self.table.item(selected_row, 0).text()

        dialog = WithdrawalDialog(quantidade_atual, valor_unitario, produto, self)
        if dialog.exec_() == QDialog.Accepted:
            quantidade_retirada = int(dialog.quantity_input.text())
            valor_retirada = quantidade_retirada * valor_unitario
            if quantidade_retirada <= quantidade_atual:  # Verifica se a quantidade de retirada é menor ou igual à quantidade atual em estoque
               data = dialog.date_edit.date().toString("dd/MM/yyyy")
               responsavel = dialog.responsavel_input.text()

               c.execute("UPDATE produtos SET quantidade=quantidade-?, valor_total=valor_total-? WHERE nome=?", (quantidade_retirada, valor_retirada, produto))
               c.execute("INSERT INTO retiradas VALUES (?, ?, ?, ?, ?, ?)", (produto, quantidade_retirada, valor_unitario, valor_retirada, data, responsavel))
               conn.commit()

               QMessageBox.information(self, "Sucesso", "Retirada registrada com sucesso!")
               self.load_products()
               self.update_total_value()
            else:
               QMessageBox.warning(self, "Erro", "A quantidade de retirada é maior do que a quantidade disponível em estoque.")
               

    def open_withdrawal_details(self, produto):
        c.execute("SELECT quantidade, valor_unitario, valor_total, data, responsavel FROM retiradas WHERE produto=?", (produto,))
        retiradas = c.fetchall()

        details_dialog = QDialog(self)
        details_dialog.setWindowTitle("Detalhes da Retirada")
        details_dialog.setFixedSize(655, 500)

        details_layout = QVBoxLayout(details_dialog)

        table = QTableWidget(details_dialog)
        table.setRowCount(len(retiradas))
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Quantidade"," Valor Unitario", "Valor Total", "Data", "Responsável"])
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        for row_number, row_data in enumerate(retiradas):
            for column_number, data in enumerate(row_data):
                item = QTableWidgetItem(str(data))
                table.setItem(row_number, column_number, item)

        details_layout.addWidget(table)

        details_dialog.exec_()
        
    def open_entry_dialog(self):
        selected_items = self.table.selectedItems()
        if len(selected_items) == 0:
            QMessageBox.warning(self, "Erro", "Selecione um item para registrar a entrada.")
            return

        selected_row = selected_items[0].row()
        quantidade_maxima = int(self.table.item(selected_row, 2).text())
        valor_unitario = float(self.table.item(selected_row, 7).text())
        produto = self.table.item(selected_row, 0).text()

        dialog = EntryDialog(quantidade_maxima, valor_unitario, produto, self)
        if dialog.exec_() == QDialog.Accepted:
            quantidade_entrada = int(dialog.quantity_input.text())
            valor_entrada = quantidade_entrada * valor_unitario
            data = dialog.date_edit.date().toString("dd/MM/yyyy")
            responsavel = dialog.responsavel_input.text()

            c.execute("UPDATE produtos SET quantidade=quantidade+?, valor_total=valor_total+? WHERE nome=?", (quantidade_entrada, valor_entrada, produto))
            c.execute("INSERT INTO entradas VALUES (?, ?, ?, ?, ?, ?)", (produto, quantidade_entrada, valor_unitario,  valor_entrada, data, responsavel))
            conn.commit()

            QMessageBox.information(self, "Sucesso", "Entrada registrada com sucesso!")
            self.load_products()
            self.update_total_value()
        
    def open_entry_details(self, produto):
        c.execute("SELECT quantidade, valor_unitario, valor_total, data, responsavel FROM entradas WHERE produto=?", (produto,))
        entradas = c.fetchall()

        details_dialog = QDialog(self)
        details_dialog.setWindowTitle("Detalhes da entrada")
        details_dialog.setFixedSize(655, 500)

        details_layout = QVBoxLayout(details_dialog)

        table = QTableWidget(details_dialog)
        table.setRowCount(len(entradas))
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Quantidade", "Valor Unitario", "Valor Total", "Data", "Responsável"])
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        for row_number, row_data in enumerate(entradas):
            for column_number, data in enumerate(row_data):
                item = QTableWidgetItem(str(data))
                table.setItem(row_number, column_number, item)

        details_layout.addWidget(table)

        details_dialog.exec_() 

    def generate_report(self):
            current_directory = os.getcwd()
            report_directory = os.path.join(current_directory, "Relatório e Verificação")
            os.makedirs(report_directory, exist_ok=True)  # Cria a pasta se não existir      
            file_path = os.path.join(report_directory, "Relatório do Almoxarifado.xlsx")
            workbook = openpyxl.Workbook()

            # Planilha de Produtos
            products_sheet = workbook.active
            products_sheet.title = "Itens"


# Ajustar a altura da primeira linha
            products_sheet.row_dimensions[1].height = 80
# Carregar a imagem
            img1 = Image('imagem.jpeg')
            img2 = Image('imagem.jpeg') # Substitua 'nome_da_imagem.png' pelo nome real da sua imagem

# Redimensionar a imagem, se necessário
            img1.width = 105
            img1.height = 105
            img2.width = 112
            img2.height = 105

# Adicionar a imagem à planilha
            products_sheet.add_image(img1, 'A1')
            products_sheet.add_image(img2, 'G1')
            

            # Definir estilo para o cabeçalho
            header_style = openpyxl.styles.NamedStyle(name="header_style")
            header_style.alignment = Alignment(horizontal="center", vertical="center")
            # Preencher cabeçalho
            products_sheet.merge_cells("A1:G1")
            header_cell = products_sheet.cell(row=1, column=1, value="Almoxarifado Caps Paulo da Portela - Relatório de Itens - Prefeitura do Rio de Janeiro")
            header_cell.style = header_style
            
# No método generate_report
            c.execute("SELECT SUM(valor_total) FROM produtos")
            total_value = c.fetchone()[0]
            if total_value is None:
                total_value = 0

# Adicionar o texto "Valor Total do Almoxarifado:" à célula B2
            total_value_text_cell = products_sheet.cell(row=2, column=2, value="T. Almoxarifado:")
            total_value_text_cell.style = header_style

# Adicionar o valor total do almoxarifado à célula C2
            total_value_cell = products_sheet.cell(row=2, column=3, value=total_value)

# Aplicar formatação de moeda à célula C2
            total_value_cell.number_format = '"R$"#,##0.00'            

            # Preencher nomes das colunas
            product_columns = ["Nome", "Categoria", "Quantidade", "Quantidade Mínima", "Vencimento", "Data do Primeiro Registro", "Valor Unitário", "Valor Total"]
            for col, column_name in enumerate(product_columns, start=1):
                cell = products_sheet.cell(row=3, column=col, value=column_name)
                cell.style = header_style

            # Preencher dados dos produtos
            c.execute("SELECT nome, categoria, quantidade, quantidade_minima, vencimento, data_registro, valor_unitario, valor_unitario * quantidade FROM produtos")
            products = c.fetchall()
            for row, product in enumerate(products, start=4):
                for col, data in enumerate(product, start=1):
                    cell = products_sheet.cell(row=row, column=col, value=data)

            # Ajustar largura das colunas
            for column in range(1, len(product_columns) + 1):
                column_letter = get_column_letter(column)
                products_sheet.column_dimensions[column_letter].width = 15

            # Planilha de Retiradas
            withdrawals_sheet = workbook.create_sheet(title="Retiradas")
            
            # Ajustar a altura da primeira linha
            withdrawals_sheet.row_dimensions[1].height = 80
# Carregar a imagem
            img1 = Image('imagem.jpeg')
            img2 = Image('imagem.jpeg') # Substitua 'nome_da_imagem.png' pelo nome real da sua imagem

# Redimensionar a imagem, se necessário
            img1.width = 95
            img1.height = 105
            img2.width = 112
            img2.height = 105

# Adicionar a imagem à planilha
            withdrawals_sheet.add_image(img1, 'A1')
            withdrawals_sheet.add_image(img2, 'G1')

            # Preencher cabeçalho
            withdrawals_sheet.merge_cells("A1:G1")
            header_cell = withdrawals_sheet.cell(row=1, column=1, value="Almoxarifado Caps Paulo da Portela - Relatório de Retiradas - Prefeitura do Rio de Janeiro")
            header_cell.style = header_style

            # Preencher nomes das colunas
            withdrawal_columns = ["Produto", "Quantidade","Valor Unitario", "Valor Total", "Data", "Responsável"]
            for col, column_name in enumerate(withdrawal_columns, start=1):
                cell = withdrawals_sheet.cell(row=3, column=col, value=column_name)
                cell.style = header_style

            # Preencher dados das retiradas
            c.execute("SELECT * FROM retiradas")
            withdrawals = c.fetchall()
            for row, withdrawal in enumerate(withdrawals, start=4):
                for col, data in enumerate(withdrawal, start=1):
                    cell = withdrawals_sheet.cell(row=row, column=col, value=data)

            # Ajustar largura das colunas
            for column in range(1, len(withdrawal_columns) + 1):
                column_letter = get_column_letter(column)
                withdrawals_sheet.column_dimensions[column_letter].width = 15

            # Planilha de Entradas
            entries_sheet = workbook.create_sheet(title="Entradas")
            
                        # Ajustar a altura da primeira linha
            entries_sheet.row_dimensions[1].height = 80
# Carregar a imagem
            img1 = Image('imagem.jpeg')
            img2 = Image('imagem.jpeg') # Substitua 'nome_da_imagem.png' pelo nome real da sua imagem

# Redimensionar a imagem, se necessário
            img1.width = 95
            img1.height = 105
            img2.width = 112
            img2.height = 105

# Adicionar a imagem à planilha
            entries_sheet.add_image(img1, 'A1')
            entries_sheet.add_image(img2, 'G1')
            

            # Preencher cabeçalho
            entries_sheet.merge_cells("A1:G1")
            header_cell = entries_sheet.cell(row=1, column=1, value="Almoxarifado Caps Paulo da Portela - Relatório de Entradas - Prefeitura do Rio de Janeiro")
            header_cell.style = header_style

            # Preencher nomes das colunas
            entry_columns = ["Produto", "Quantidade", "Valor Unitario", "Valor Total", "Data", "Responsável"]
            for col, column_name in enumerate(entry_columns, start=1):
                cell = entries_sheet.cell(row=3, column=col, value=column_name)
                cell.style = header_style

            # Preencher dados das entradas
            c.execute("SELECT * FROM entradas")
            entries = c.fetchall()
            for row, entry in enumerate(entries, start=4):
                for col, data in enumerate(entry, start=1):
                    cell = entries_sheet.cell(row=row, column=col, value=data)

            # Ajustar largura das colunas
            for column in range(1, len(entry_columns) + 1):
                column_letter = get_column_letter(column)
                entries_sheet.column_dimensions[column_letter].width = 15


            # Salvar o arquivo
            workbook.save(file_path)
            QMessageBox.information(self, "Relatório Gerado", "O relatório foi gerado com sucesso!")
     

    def clear_input_fields(self):
        self.name_input.clear()
        self.quantity_input.clear()
        self.min_quantity_input.clear()
        self.day_input.clear()
        self.month_input.clear()
        self.year_input.clear()
        self.register_day_input.clear()
        self.register_month_input.clear()
        self.register_year_input.clear()
        self.unit_value_input.clear()

if __name__ == "__main__":
    app = QApplication(sys.argv)

    main_window = MainWindow()
    main_window.show()

    sys.exit(app.exec_())
